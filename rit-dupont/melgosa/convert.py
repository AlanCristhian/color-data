import json
import openpyxl
import yaml

wb = openpyxl.load_workbook("RIT-DuPont.xlsx")

ws = wb.worksheets[0]
dv = [ws[f"A{k}"].value for k in range(3, 315)]

white = [ws["B3"].value, ws["C3"].value, ws["D3"].value]

pairs = [
    [
        [ws[f"E{k}"].value, ws[f"F{k}"].value, ws[f"G{k}"].value],
        [ws[f"H{k}"].value, ws[f"I{k}"].value, ws[f"J{k}"].value],
    ]
    for k in range(3, 315)
]

d = {"reference_white": white, "dv": dv, "pairs": pairs}
with open("rit-dupont.yaml", "w") as f:
    yaml.dump(d, f)

with open("rit-dupont.json", "w") as f:
    json.dump(d, f, indent=2)
